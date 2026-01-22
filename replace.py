#!/usr/bin/env python3

import csv
import json
from datetime import datetime
from fuz import fuzzy_match
from openpyxl import Workbook
from openpyxl.styles import PatternFill


def load_links_csv(csv_path):
    id_to_idx = {}
    all_rows = []
    
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for idx, row in enumerate(reader):
            row_name = row['name'].strip()
            row_id = row['id'].strip()
            row_link = row['link'].strip() if row['link'] else ''
            
            row_tuple = (row_name, row_id, row_link)
            all_rows.append(row_tuple)
            
            if row_id not in id_to_idx:
                id_to_idx[row_id] = idx
    
    return id_to_idx, all_rows


def find_partial_id_match(prop_id, all_rows, used_csv_indices):
    for idx, row in enumerate(all_rows):
        if idx in used_csv_indices:
            continue
        csv_id = row[1]
        if csv_id in prop_id:
            return idx, row
    return None, None


def find_best_fuzzy_match(prop_name, all_rows, used_csv_indices, min_score=0.4):
    best_idx = None
    best_row = None
    best_score = 0.0
    
    for idx, row in enumerate(all_rows):
        if idx in used_csv_indices:
            continue
        row_name = row[0]
        score = fuzzy_match(prop_name, row_name)
        if score > best_score:
            best_score = score
            best_idx = idx
            best_row = row
    
    if best_score >= min_score:
        return best_idx, best_row, best_score
    return None, None, 0.0


def write_excel(excel_path, found_by_id, found_by_id_partial, found_by_fuzzy, not_found_list):
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    
    ws['A1'] = 'Pozi'
    ws['B1'] = 'WMIS'
    ws['C1'] = 'Matched'
    
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    
    all_rows = []
    
    for pozi_name, csv_name in found_by_id:
        all_rows.append((pozi_name, csv_name, 'Yes'))
    
    for pozi_name, csv_name in found_by_id_partial:
        all_rows.append((pozi_name, csv_name, 'Yes'))
    
    for pozi_name, csv_name in found_by_fuzzy:
        all_rows.append((pozi_name, csv_name, 'No'))
    
    for pozi_name in not_found_list:
        all_rows.append((pozi_name, '', 'No'))
    
    all_rows.sort(key=lambda x: (0 if x[2] == 'Yes' else 1, x[0].lower()))
    
    row_num = 2
    for pozi_name, csv_name, matched in all_rows:
        ws.cell(row=row_num, column=1, value=pozi_name)
        ws.cell(row=row_num, column=2, value=csv_name)
        ws.cell(row=row_num, column=3, value=matched)
        if matched == 'Yes':
            ws.cell(row=row_num, column=1).fill = green_fill
            ws.cell(row=row_num, column=2).fill = green_fill
            ws.cell(row=row_num, column=3).fill = green_fill
        row_num += 1
    
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 10
    
    wb.save(excel_path)


def process_geojson(source_path, output_path, id_to_idx, all_rows, log_path, excel_path):
    with open(source_path, 'r', encoding='utf-8') as f:
        geojson_data = json.load(f)
    
    features = geojson_data.get('features', [])
    
    found_by_id = []
    found_by_id_partial = []
    found_by_fuzzy = []
    not_found_list = []
    
    excel_by_id = []
    excel_by_id_partial = []
    excel_by_fuzzy = []
    excel_not_found = []
    
    used_csv_indices = set()
    matched_feature_indices = set()
    
    for feat_idx, feature in enumerate(features):
        properties = feature.get('properties', {})
        if properties is None:
            continue
        prop_id = properties.get('id')
        if prop_id is None:
            continue
        
        prop_id_trimmed = str(prop_id).strip()
        prop_name_str = str(properties.get('name')) if properties.get('name') else ''
        
        if prop_id_trimmed in id_to_idx:
            csv_idx = id_to_idx[prop_id_trimmed]
            if csv_idx not in used_csv_indices:
                csv_row = all_rows[csv_idx]
                properties['source'] = csv_row[2]
                found_by_id.append((prop_name_str, prop_id_trimmed))
                excel_by_id.append((prop_name_str, csv_row[0]))
                used_csv_indices.add(csv_idx)
                matched_feature_indices.add(feat_idx)
    
    for feat_idx, feature in enumerate(features):
        if feat_idx in matched_feature_indices:
            continue
        properties = feature.get('properties', {})
        if properties is None:
            continue
        prop_id = properties.get('id')
        if prop_id is None:
            continue
        
        prop_id_trimmed = str(prop_id).strip()
        prop_name_str = str(properties.get('name')) if properties.get('name') else ''
        
        csv_idx, csv_row = find_partial_id_match(prop_id_trimmed, all_rows, used_csv_indices)
        if csv_row:
            properties['source'] = csv_row[2]
            found_by_id_partial.append((prop_name_str, prop_id_trimmed, csv_row[0], csv_row[1]))
            excel_by_id_partial.append((prop_name_str, csv_row[0]))
            used_csv_indices.add(csv_idx)
            matched_feature_indices.add(feat_idx)
    
    for feat_idx, feature in enumerate(features):
        if feat_idx in matched_feature_indices:
            continue
        properties = feature.get('properties', {})
        if properties is None:
            continue
        prop_id = properties.get('id')
        if prop_id is None:
            continue
        
        prop_id_trimmed = str(prop_id).strip()
        prop_name_str = str(properties.get('name')) if properties.get('name') else ''
        
        best_idx, best_row, best_score = find_best_fuzzy_match(
            prop_name_str, all_rows, used_csv_indices, min_score=0.4
        )
        if best_row:
            properties['source'] = best_row[2]
            found_by_fuzzy.append((prop_name_str, prop_id_trimmed, best_row[0], best_row[1], best_score))
            excel_by_fuzzy.append((prop_name_str, best_row[0]))
            used_csv_indices.add(best_idx)
            matched_feature_indices.add(feat_idx)
        else:
            not_found_list.append((prop_name_str, prop_id_trimmed))
            excel_not_found.append(prop_name_str)
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(geojson_data, f, indent=2)
    
    with open(log_path, 'w', encoding='utf-8') as f:
        f.write(f"replace.py Log - Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("=" * 80 + "\n\n")
        
        f.write("(1) FOUND BY ID\n")
        f.write("-" * 40 + "\n")
        for name, pid in found_by_id:
            f.write(f"  Name: {name}, ID: {pid}\n")
        f.write(f"\nTotal found by ID: {len(found_by_id)}\n\n")
        
        f.write("(2) FOUND BY ID (partial - CSV id is substring of geojson id)\n")
        f.write("-" * 40 + "\n")
        for src_name, src_id, csv_name, csv_id in found_by_id_partial:
            f.write(f"  source.geojson -> Name: {src_name}, ID: {src_id}\n")
            f.write(f"  links.csv      -> Name: {csv_name}, ID: {csv_id}\n")
            f.write("\n")
        f.write(f"Total found by ID (partial): {len(found_by_id_partial)}\n\n")
        
        f.write("(3) FOUND BY NAME (fuzzy match, score >= 0.4)\n")
        f.write("-" * 40 + "\n")
        for src_name, src_id, csv_name, csv_id, score in found_by_fuzzy:
            f.write(f"  source.geojson -> Name: {src_name}, ID: {src_id}\n")
            f.write(f"  links.csv      -> Name: {csv_name}, ID: {csv_id}\n")
            f.write(f"  Fuzzy score: {score}\n")
            f.write("\n")
        f.write(f"Total found by fuzzy: {len(found_by_fuzzy)}\n\n")
        
        f.write("(4) NOT FOUND\n")
        f.write("-" * 40 + "\n")
        for name, pid in not_found_list:
            f.write(f"  source.geojson -> Name: {name}, ID: {pid}\n")
        f.write(f"\nTotal not found: {len(not_found_list)}\n")
    
    write_excel(excel_path, excel_by_id, excel_by_id_partial, excel_by_fuzzy, excel_not_found)
    
    return len(found_by_id), len(found_by_id_partial), len(found_by_fuzzy), len(not_found_list)


def main():
    csv_path = 'links.csv'
    source_path = 'source.geojson'
    output_path = 'River Gauges.geojson'
    log_path = 'replace.log'
    excel_path = 'replace.xlsx'
    
    print(f"Loading links from {csv_path}...")
    id_to_idx, all_rows = load_links_csv(csv_path)
    print(f"Loaded {len(all_rows)} rows ({len(id_to_idx)} unique IDs).")
    
    print(f"Processing {source_path}...")
    id_count, id_partial_count, fuzzy_count, not_found_count = process_geojson(
        source_path, output_path, id_to_idx, all_rows, log_path, excel_path
    )
    
    print(f"\nDone!")
    print(f"  - Output written to: {output_path}")
    print(f"  - Log written to: {log_path}")
    print(f"  - Excel written to: {excel_path}")
    print(f"  - Found by ID: {id_count}")
    print(f"  - Found by ID (partial): {id_partial_count}")
    print(f"  - Found by Fuzzy: {fuzzy_count}")
    print(f"  - Not found: {not_found_count}")


if __name__ == '__main__':
    main()
